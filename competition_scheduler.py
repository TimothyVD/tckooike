#!/usr/bin/env python3
"""
Poule Competition Scheduler
============================
Loads team availabilities from a Google Forms CSV export and terrain slot
definitions, then schedules all round-robin matches optimally using CP-SAT.

Constraints enforced
---------------------
- Both teams must be available at the assigned slot.
- Each (slot, terrain) pair is used by at most one match.
- Each team plays at most one match per day.
- Every match is scheduled at most once.

Objective
----------
Maximise the number of scheduled matches (minimise unscheduled ones).

Installation
-------------
    pip install ortools pandas openpyxl

Google Forms setup (recommended)
----------------------------------
Create a form with:
  - Short answer question: "Team name"
  - Checkbox grid question: rows = dates, columns = time slots
    → Export responses as CSV (Google Sheets → File → Download → CSV)

Usage
------
    python competition_scheduler.py                        # runs built-in demo
    python competition_scheduler.py --teams teams.csv \\
           --avail team_availabilities.csv \\
           --slots terrain_slots.csv \\
           --output schedule.xlsx
"""

from __future__ import annotations

import argparse
import random
import re
import site_builder
from itertools import combinations
from pathlib import Path

import pandas as pd
from ortools.sat.python import cp_model


# ── Constants ──────────────────────────────────────────────────────────────────

SLOT_DURATION_MIN = 90  # minutes per match slot (1h30)

# Truthy strings recognised in the availability CSV
_TRUTHY = {"true", "checked", "yes", "oui", "1", "x", "✓"}


# ── Input loading ──────────────────────────────────────────────────────────────

def load_team_availabilities(csv_path: str) -> dict[str, set[str]]:
    """
    Load team availabilities from a Google Forms checkbox-grid CSV export.

    Expected CSV layout
    --------------------
    Two supported formats:

    Format A — checkbox grid (one column per slot):
        Team,2026-05-04 18:00,2026-05-04 19:30,2026-05-11 18:00,...
        Alpha,TRUE,,TRUE,...
        Bravo,,TRUE,TRUE,...

    Format B — multi-select (selected slots comma-separated in one cell):
        Team,Available slots
        Alpha,"2026-05-04 18:00, 2026-05-11 18:00"
        Bravo,"2026-05-04 19:30"

    The column named "Team" (case-insensitive) holds team names; every other
    column is treated as a slot label or a multi-select availability field.

    Slot labels must match those used in the terrain slots definition
    (format: "YYYY-MM-DD HH:MM").

    Returns
    --------
    {team_name: {slot_label, ...}}
    """
    df = pd.read_csv(csv_path, dtype=str).fillna("")

    # Find the team-name column (case-insensitive)
    team_col = next(
        (c for c in df.columns if c.strip().lower() == "team"), df.columns[0]
    )
    other_cols = [c for c in df.columns if c != team_col]

    availabilities: dict[str, set[str]] = {}

    for _, row in df.iterrows():
        team = row[team_col].strip()
        if not team:
            continue

        available: set[str] = set()

        if len(other_cols) == 1:
            # Format B: comma-separated slot list in a single cell
            raw = row[other_cols[0]]
            for slot in raw.split(","):
                s = slot.strip()
                if s:
                    available.add(s)
        else:
            # Format A: one column per slot, truthy value = available
            for col in other_cols:
                if row[col].strip().lower() in _TRUTHY:
                    available.add(col.strip())

        availabilities[team] = available

    return availabilities


def load_terrain_slots(csv_path: str) -> list[dict]:
    """
    Load terrain slot definitions from a CSV file.

    Expected CSV columns
    ----------------------
        date        YYYY-MM-DD
        time        HH:MM
        terrain_id  integer (1, 2, 3, …)

    Example
    --------
        date,time,terrain_id
        2026-05-04,18:00,1
        2026-05-04,18:00,2
        2026-05-04,19:30,1
        ...

    Returns
    --------
    List of dicts with keys: slot ("YYYY-MM-DD HH:MM"), date, terrain ("T1"…).
    """
    df = pd.read_csv(csv_path, dtype=str).fillna("")
    df["slot"] = df["date"].str.strip() + " " + df["time"].str.strip()
    df["terrain"] = "T" + df["terrain_id"].str.strip()
    return df[["slot", "date", "terrain"]].to_dict("records")


def generate_terrain_slots(
    dates: list[str],
    times: list[str],
    n_terrains: int = 4,
    terrain_overrides: dict[str, int] | None = None,
) -> list[dict]:
    """
    Programmatically generate terrain slots.

    Parameters
    -----------
    dates           List of date strings, e.g. ["2026-05-04", "2026-05-11"]
    times           Time strings per day, e.g. ["18:00", "19:30", "21:00"]
    n_terrains      Default number of terrains per slot
    terrain_overrides  Per-date override: {"2026-05-04": 2} → only 2 terrains
    """
    overrides = terrain_overrides or {}
    slots = []
    for date in dates:
        n = overrides.get(date, n_terrains)
        for time in times:
            for t in range(1, n + 1):
                slots.append(
                    {"slot": f"{date} {time}", "date": date, "terrain": f"T{t}"}
                )
    return slots


def _make_team_name(player_1: str, player_2: str) -> str:
    """
    Build a display team name from two player name strings.

    Format: "<FirstName> <LastInitial> & <FirstName> <LastInitial>"
    Example: "Alice Dupont" + "Bob Martin"  →  "Alice D & Bob M"

    Falls back gracefully when names have no surname or are blank.
    """
    def _fmt(name: str) -> str:
        parts = name.strip().split()
        if not parts:
            return ""
        first = parts[0]
        initial = parts[-1][0].upper() + "." if len(parts) > 1 else ""
        return f"{first} {initial}".strip()

    a = _fmt(player_1)
    b = _fmt(player_2)
    if a and b:
        return f"{a} & {b}"
    return a or b


def load_teams(csv_path: str) -> tuple[dict[str, list[str]], pd.DataFrame]:
    """
    Load team names, poule assignments, and optional player metadata from a CSV.
    Column order is flexible — columns are matched by name (case-insensitive).

    Required column
    ----------------
        team               Team name (optional when player_1 and player_2 are
                           both present — auto-generated as "FirstName L & FirstName L")

    Optional columns (any subset, in any order)
    ---------------------------------------------
        poule              Poule/group identifier (defaults to "A" if absent)
        player_1           Name of first player
        player_2           Name of second player
        ranking_player_1   Ranking / level of player 1
        ranking_player_2   Ranking / level of player 2
        tel_player_1       Phone number of player 1
        tel_player_2       Phone number of player 2

    Returns
    --------
    (poules, team_info_df)
      poules        {poule_name: [team1, team2, …]}
      team_info_df  Full DataFrame with all columns, normalised column names,
                    sorted by Poule then Team — for writing to the Teams sheet.
    """
    df = pd.read_csv(csv_path, dtype=str).fillna("")

    # Normalise column names: strip whitespace, lower-case for lookup
    col_map = {c.strip().lower(): c for c in df.columns}

    def _find(name: str):
        return col_map.get(name)

    team_col  = _find("team")
    if team_col is None:
        team_col = df.columns[0]  # fall back to first column
    poule_col = _find("poule")

    # Canonical display names for known extra columns (order defines sheet column order)
    _KNOWN_EXTRAS = [
        ("poule",            "Poule"),
        ("player_1",         "Player 1"),
        ("player_2",         "Player 2"),
        ("ranking_player_1", "Ranking P1"),
        ("ranking_player_2", "Ranking P2"),
        ("tel_player_1",     "Tel P1"),
        ("tel_player_2",     "Tel P2"),
    ]

    poules: dict[str, list[str]] = {}
    info_rows = []

    p1_col = _find("player_1")
    p2_col = _find("player_2")

    for _, row in df.iterrows():
        p1 = row[p1_col].strip() if p1_col else ""
        p2 = row[p2_col].strip() if p2_col else ""

        # Auto-generate team name from players when both are present;
        # fall back to the explicit Team column value otherwise.
        if p1 and p2:
            team = _make_team_name(p1, p2)
        else:
            team = row[team_col].strip() if team_col else ""
        if not team:
            continue
        poule = row[poule_col].strip() if poule_col else "A"
        if not poule:
            poule = "A"
        poules.setdefault(poule, []).append(team)

        info = {"Team": team}
        for key, display in _KNOWN_EXTRAS:
            src = _find(key)
            info[display] = row[src].strip() if src else ""
        info_rows.append(info)

    # Build display DataFrame with canonical column order, drop empty columns
    display_cols = ["Team"] + [display for _, display in _KNOWN_EXTRAS]
    team_info_df = pd.DataFrame(info_rows, columns=display_cols)
    # Remove columns that are entirely empty (not provided in the source CSV)
    team_info_df = team_info_df.loc[:, (team_info_df != "").any(axis=0)]
    team_info_df = team_info_df.sort_values(["Poule", "Team"] if "Poule" in team_info_df.columns else ["Team"]).reset_index(drop=True)

    return poules, team_info_df




# ── Core scheduler ─────────────────────────────────────────────────────────────

def schedule(
    poules: dict[str, list[str]],
    team_avail: dict[str, set[str]],
    terrain_slots: list[dict],
    time_limit_s: int = 60,
    verbose: bool = True,
) -> tuple[list[dict], list[tuple[str, str, str]]]:
    """
    Assign all intra-poule round-robin matches to (slot, terrain) pairs using CP-SAT.
    All poules share the same terrain slots (they compete for the same venue).

    Parameters
    -----------
    poules          {poule_name: [team, …]} — teams grouped by poule.
    team_avail      {team_name: {slot_label, …}} — availability per team.
    terrain_slots   List of slot dicts from generate_terrain_slots / load_terrain_slots.
    time_limit_s    Maximum solver wall-clock time in seconds.
    verbose         Print a summary after solving.

    Returns
    --------
    (scheduled, unscheduled)
      scheduled   — list of {"poule": "A", "match": (A, B), "slot": "…", "terrain": "T1"}
      unscheduled — list of (poule, A, B) tuples that could not be placed
    """
    # Build flat match list with poule tag (only intra-poule round-robin)
    matches: list[tuple[str, str, str]] = [
        (poule, ta, tb)
        for poule, teams in poules.items()
        for ta, tb in combinations(teams, 2)
    ]
    all_teams = [t for ts in poules.values() for t in ts]

    # De-duplicate slot labels while preserving order
    slot_labels: list[str] = list(dict.fromkeys(s["slot"] for s in terrain_slots))
    date_of_slot: dict[str, str] = {s["slot"]: s["date"] for s in terrain_slots}
    # All actual (slot, terrain) pairs that the venue provides
    st_pairs: list[tuple[str, str]] = [(s["slot"], s["terrain"]) for s in terrain_slots]

    model = cp_model.CpModel()

    # ── Decision variables ────────────────────────────────────────────────────
    # x[(match_idx, st_idx)] = 1  iff  match is played at that (slot, terrain)
    # Variables are only created when both teams are available — prunes the model.
    x: dict[tuple[int, int], cp_model.IntVar] = {}

    for m_idx, (_, ta, tb) in enumerate(matches):
        avail_a = team_avail.get(ta, set())
        avail_b = team_avail.get(tb, set())
        for st_idx, (slot, _terrain) in enumerate(st_pairs):
            if slot in avail_a and slot in avail_b:
                x[(m_idx, st_idx)] = model.new_bool_var(f"x_m{m_idx}_st{st_idx}")

    # ── Constraints ───────────────────────────────────────────────────────────

    # 1. Each match is played at most once
    for m_idx in range(len(matches)):
        match_vars = [x[(m_idx, st)] for st in range(len(st_pairs)) if (m_idx, st) in x]
        if match_vars:
            model.add(sum(match_vars) <= 1)

    # 2. Each (slot, terrain) pair hosts at most one match
    for st_idx in range(len(st_pairs)):
        slot_vars = [x[(m, st_idx)] for m in range(len(matches)) if (m, st_idx) in x]
        if slot_vars:
            model.add(sum(slot_vars) <= 1)

    # 3. Each team plays at most one match per calendar day
    all_dates = list(dict.fromkeys(s["date"] for s in terrain_slots))
    for team in all_teams:
        for date in all_dates:
            st_on_date = [
                st_idx
                for st_idx, (slot, _) in enumerate(st_pairs)
                if date_of_slot[slot] == date
            ]
            team_day_vars = [
                x[(m_idx, st_idx)]
                for m_idx, (_, ta, tb) in enumerate(matches)
                if ta == team or tb == team
                for st_idx in st_on_date
                if (m_idx, st_idx) in x
            ]
            if team_day_vars:
                model.add(sum(team_day_vars) <= 1)

    # ── Objective: schedule as many matches as possible ───────────────────────
    model.maximize(sum(x.values()))

    # ── Solve ─────────────────────────────────────────────────────────────────
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = float(time_limit_s)
    solver.parameters.log_search_progress = False
    status = solver.solve(model)

    scheduled: list[dict] = []
    scheduled_matches: set[int] = set()

    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        for m_idx, (poule, ta, tb) in enumerate(matches):
            for st_idx, (slot, terrain) in enumerate(st_pairs):
                if (m_idx, st_idx) in x and solver.value(x[(m_idx, st_idx)]):
                    scheduled.append(
                        {"poule": poule, "match": (ta, tb), "slot": slot, "terrain": terrain}
                    )
                    scheduled_matches.add(m_idx)
                    break

    unscheduled = [
        (poule, ta, tb)
        for i, (poule, ta, tb) in enumerate(matches)
        if i not in scheduled_matches
    ]

    if verbose:
        status_name = solver.status_name(status)
        total = len(matches)
        n_sch = len(scheduled)
        n_slots = len(slot_labels)
        n_slot_terrain = len(st_pairs)
        print(f"\n{'='*60}")
        print(f"  Poules:         {len(poules)}")
        print(f"  Total matches:  {total}")
        print(f"  Scheduled:      {n_sch}  ({100*n_sch/total:.0f}%)")
        print(f"  Unscheduled:    {len(unscheduled)}")
        print(f"  Venue slots:    {n_slot_terrain}  ({n_slots} time slots x terrains)")
        print(f"  Solver status:  {status_name}")
        for poule_name, teams in poules.items():
            p_total = len(list(combinations(teams, 2)))
            p_sch   = sum(1 for s in scheduled if s["poule"] == poule_name)
            print(f"  Poule {poule_name:6s}:   {p_sch}/{p_total} matches  ({', '.join(teams)})")
        print(f"{'='*60}\n")

        if unscheduled:
            print("Unscheduled matches (no common availability):")
            for poule, ta, tb in unscheduled:
                print(f"  [{poule}]  {ta}  vs  {tb}")
            print()

    return scheduled, unscheduled


# ── Output helpers ─────────────────────────────────────────────────────────────

def to_dataframes(
    scheduled: list[dict],
    unscheduled: list[tuple[str, str, str]],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Convert results to sorted DataFrames."""
    if scheduled:
        df_sched = pd.DataFrame(
            [
                {
                    "Poule": s["poule"],
                    "Date": s["slot"].split()[0],
                    "Time": s["slot"].split()[1],
                    "Terrain": s["terrain"],
                    "Team A": s["match"][0],
                    "Team B": s["match"][1],
                }
                for s in scheduled
            ]
        ).sort_values(["Poule", "Date", "Time", "Terrain"]).reset_index(drop=True)
    else:
        df_sched = pd.DataFrame(columns=["Poule", "Date", "Time", "Terrain", "Team A", "Team B"])

    df_unsched = (
        pd.DataFrame([{"Poule": p, "Team A": a, "Team B": b} for p, a, b in unscheduled])
        if unscheduled
        else pd.DataFrame(columns=["Poule", "Team A", "Team B"])
    )

    return df_sched, df_unsched


def export_schedule_overview_pdf(
    df_sched: pd.DataFrame,
    path: str,
    club_name: str = "Tennis Club",
    season: str = "",
) -> None:
    """
    Generate a compact planning overview PDF:
    all matches from all poules combined, sorted by date then time,
    displayed in a single chronological table.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError:
        print("reportlab not installed — skipping overview PDF.  pip install reportlab")
        return

    if df_sched.empty:
        print("No schedule data — skipping overview PDF.")
        return

    PAGE_W, PAGE_H = A4
    MARGIN = 2 * cm

    CLAY        = colors.HexColor("#C1440E")
    CLAY_LIGHT  = colors.HexColor("#F2E0D0")
    DARK        = colors.HexColor("#1C1C1C")
    MID_GREY    = colors.HexColor("#6D6D6D")
    WHITE       = colors.white
    DAY_BG      = colors.HexColor("#F7F0EC")   # subtle warm tint for date group headers

    styles = getSampleStyleSheet()

    S_TITLE = ParagraphStyle(
        "Title2", parent=styles["Title"],
        fontSize=26, textColor=CLAY, spaceAfter=4,
        fontName="Helvetica-Bold", alignment=TA_CENTER,
    )
    S_SUBTITLE = ParagraphStyle(
        "Subtitle2", parent=styles["Normal"],
        fontSize=12, textColor=MID_GREY, spaceAfter=2,
        fontName="Helvetica", alignment=TA_CENTER,
    )

    # Sort all matches by date then time
    df_sorted = df_sched.sort_values(["Date", "Time", "Terrain"]).reset_index(drop=True)

    # Decide columns: include Poule only if multiple poules exist
    has_poule = "Poule" in df_sorted.columns and df_sorted["Poule"].nunique() > 1

    # Column widths
    if has_poule:
        col_hdrs = ["Date", "Time", "Terrain", "Poule", "Team A", "vs", "Team B"]
        date_w    = 2.4 * cm
        time_w    = 1.6 * cm
        terr_w    = 1.6 * cm
        poule_w   = 1.5 * cm
        vs_w      = 0.7 * cm
        team_w    = (PAGE_W - 2 * MARGIN - date_w - time_w - terr_w - poule_w - vs_w) / 2
        col_widths = [date_w, time_w, terr_w, poule_w, team_w, vs_w, team_w]
    else:
        col_hdrs = ["Date", "Time", "Terrain", "Team A", "vs", "Team B"]
        date_w    = 2.6 * cm
        time_w    = 1.8 * cm
        terr_w    = 1.8 * cm
        vs_w      = 0.8 * cm
        team_w    = (PAGE_W - 2 * MARGIN - date_w - time_w - terr_w - vs_w) / 2
        col_widths = [date_w, time_w, terr_w, team_w, vs_w, team_w]

    # Build table rows, inserting a shaded date-group separator row on date change
    table_data = [col_hdrs]
    row_styles = []   # extra TableStyle commands collected per row
    current_date = None
    data_row_idx = 1  # index 0 = header

    for _, row in df_sorted.iterrows():
        date = row["Date"]
        if date != current_date:
            current_date = date
            # Date group header row (spans all columns)
            span_row = [date] + [""] * (len(col_hdrs) - 1)
            table_data.append(span_row)
            row_styles.append(
                ("SPAN",        (0, data_row_idx), (-1, data_row_idx))
            )
            row_styles.append(
                ("BACKGROUND",  (0, data_row_idx), (-1, data_row_idx), DAY_BG)
            )
            row_styles.append(
                ("FONTNAME",    (0, data_row_idx), (-1, data_row_idx), "Helvetica-Bold")
            )
            row_styles.append(
                ("FONTSIZE",    (0, data_row_idx), (-1, data_row_idx), 9)
            )
            row_styles.append(
                ("TEXTCOLOR",   (0, data_row_idx), (-1, data_row_idx), DARK)
            )
            data_row_idx += 1

        if has_poule:
            table_data.append([
                "",  # date already shown in group header
                row["Time"], row["Terrain"], row.get("Poule", ""),
                row["Team A"], "vs", row["Team B"],
            ])
        else:
            table_data.append([
                "", row["Time"], row["Terrain"],
                row["Team A"], "vs", row["Team B"],
            ])
        data_row_idx += 1

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    base_style = [
        # Header row
        ("BACKGROUND",  (0, 0), (-1, 0),  CLAY),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  WHITE),
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0),  9),
        # Body
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("TEXTCOLOR",   (0, 1), (-1, -1), DARK),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, CLAY_LIGHT]),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ] + row_styles
    t.setStyle(TableStyle(base_style))

    story = []
    story.append(Spacer(1, 1.2 * cm))
    story.append(Paragraph(club_name, S_TITLE))
    if season:
        story.append(Paragraph(season, S_SUBTITLE))
    story.append(Paragraph("Full Planning Overview", S_SUBTITLE))
    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width="100%", thickness=2, color=CLAY))
    story.append(Spacer(1, 0.6 * cm))
    story.append(t)

    doc = SimpleDocTemplate(
        path, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=MARGIN,
        title=f"{club_name} — Full Planning Overview",
        author=club_name,
    )
    doc.build(story)
    print(f"Saved overview PDF to: {path}")


def export_pdf(
    df_sched: pd.DataFrame,
    team_info_df: pd.DataFrame | None,
    path: str,
    club_name: str = "Tennis Club",
    season: str = "",
) -> None:
    """
    Generate a tennis-club-style PDF with:
      - Cover header (club name, season)
      - Per-poule team sheet: team name, players, rankings, phone numbers
      - Per-team match schedule table
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            HRFlowable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table,
            TableStyle,
        )
    except ImportError:
        print("reportlab not installed — skipping PDF.  pip install reportlab")
        return

    PAGE_W, PAGE_H = A4
    MARGIN = 2 * cm

    # ── Colour palette (tennis / clay) ────────────────────────────────────────
    CLAY      = colors.HexColor("#C1440E")   # terracotta / header
    CLAY_LIGHT= colors.HexColor("#F2E0D0")   # soft peach for alt rows
    DARK      = colors.HexColor("#1C1C1C")
    MID_GREY  = colors.HexColor("#6D6D6D")
    WHITE     = colors.white
    GREEN_COURT = colors.HexColor("#2D6A4F") # dark green accent

    styles = getSampleStyleSheet()

    S_TITLE = ParagraphStyle(
        "ClubTitle",
        parent=styles["Title"],
        fontSize=28,
        textColor=CLAY,
        spaceAfter=4,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
    )
    S_SEASON = ParagraphStyle(
        "Season",
        parent=styles["Normal"],
        fontSize=13,
        textColor=MID_GREY,
        spaceAfter=2,
        fontName="Helvetica",
        alignment=TA_CENTER,
    )
    S_POULE_HDR = ParagraphStyle(
        "PouleHeader",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=WHITE,
        backColor=CLAY,
        spaceAfter=6,
        spaceBefore=14,
        fontName="Helvetica-Bold",
        leftIndent=-MARGIN + 2 * cm,
        rightIndent=-MARGIN + 2 * cm,
        borderPad=6,
        alignment=TA_LEFT,
    )
    S_SECTION = ParagraphStyle(
        "Section",
        parent=styles["Heading2"],
        fontSize=11,
        textColor=GREEN_COURT,
        spaceBefore=10,
        spaceAfter=4,
        fontName="Helvetica-Bold",
    )
    S_BODY = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontSize=9,
        textColor=DARK,
        fontName="Helvetica",
    )

    def _team_table(rows, col_headers):
        col_count = len(col_headers)
        col_w = (PAGE_W - 2 * MARGIN) / col_count
        data = [col_headers] + rows
        t = Table(data, colWidths=[col_w] * col_count, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  CLAY),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  WHITE),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0),  9),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, CLAY_LIGHT]),
            ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",    (0, 1), (-1, -1), 8),
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",  (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    def _schedule_table(rows, col_headers):
        # Measure the widest content in the "Round" column (index 3) and
        # "Opponent" column (index 4) to size them proportionally.
        # Columns: Date(0), Time(1), Terrain(2), Round(3), Opponent(4)
        from reportlab.pdfbase.pdfmetrics import stringWidth
        FONT_BODY, FONT_SIZE_BODY = "Helvetica", 8
        FONT_HDR,  FONT_SIZE_HDR  = "Helvetica-Bold", 9
        PADDING = 0.3 * cm  # horizontal padding per cell (each side)

        def _col_min_width(col_idx):
            all_vals = [col_headers[col_idx]] + [r[col_idx] for r in rows if r[col_idx]]
            return max(
                stringWidth(str(v),
                            FONT_HDR  if i == 0 else FONT_BODY,
                            FONT_SIZE_HDR if i == 0 else FONT_SIZE_BODY)
                for i, v in enumerate(all_vals)
            ) + 2 * PADDING

        date_w  = 2.2 * cm
        time_w  = 1.6 * cm
        terr_w  = 1.6 * cm
        round_w = max(_col_min_width(3), 2.5 * cm)
        opp_w   = max(_col_min_width(4), 2.5 * cm)
        # If everything fits, honour natural widths; otherwise fill the page.
        available = PAGE_W - 2 * MARGIN
        fixed     = date_w + time_w + terr_w
        used      = round_w + opp_w
        if fixed + used <= available:
            # Distribute any leftover space equally to round/opponent columns
            extra = (available - fixed - used) / 2
            round_w += extra
            opp_w   += extra
        else:
            # Scale round/opponent proportionally to fill available width
            scale  = (available - fixed) / used
            round_w *= scale
            opp_w   *= scale

        widths = [date_w, time_w, terr_w, round_w, opp_w]
        data = [col_headers] + rows
        t = Table(data, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, 0),  GREEN_COURT),
            ("TEXTCOLOR",   (0, 0), (-1, 0),  WHITE),
            ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",    (0, 0), (-1, 0),  9),
            ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, CLAY_LIGHT]),
            ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",    (0, 1), (-1, -1), 8),
            ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("TOPPADDING",  (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return t

    # ── Build story ───────────────────────────────────────────────────────────
    story = []

    # Cover header
    story.append(Spacer(1, 1.5 * cm))
    story.append(Paragraph(club_name, S_TITLE))
    if season:
        story.append(Paragraph(season, S_SEASON))
    story.append(Paragraph("Competition Schedule", S_SEASON))
    story.append(Spacer(1, 0.5 * cm))
    story.append(HRFlowable(width="100%", thickness=2, color=CLAY))
    story.append(Spacer(1, 0.8 * cm))

    # Determine poules to render
    if df_sched.empty and (team_info_df is None or team_info_df.empty):
        story.append(Paragraph("No schedule data available.", S_BODY))
    else:
        poule_names = sorted(
            df_sched["Poule"].unique().tolist() if not df_sched.empty else []
        )
        if team_info_df is not None and "Poule" in team_info_df.columns:
            poule_names = sorted(set(poule_names) | set(team_info_df["Poule"].unique()))

        for p_idx, poule in enumerate(poule_names):
            if p_idx > 0:
                story.append(PageBreak())

            story.append(Paragraph(f"  Poule {poule}", S_POULE_HDR))
            story.append(Spacer(1, 0.4 * cm))

            # ── Team info block ──────────────────────────────────────────────
            if team_info_df is not None and not team_info_df.empty:
                p_teams = team_info_df[team_info_df["Poule"] == poule] if "Poule" in team_info_df.columns else team_info_df

                # Determine which detail columns are present
                detail_cols = [c for c in ["Player 1", "Player 2", "Tel P1", "Tel P2"]
                               if c in p_teams.columns]
                team_hdrs = ["Team"] + detail_cols

                team_rows = []
                for _, row in p_teams.iterrows():
                    team_rows.append([row.get("Team", "")] + [row.get(c, "") for c in detail_cols])

                if team_rows:
                    story.append(Paragraph("Teams & Players", S_SECTION))
                    story.append(_team_table(team_rows, team_hdrs))
                    story.append(Spacer(1, 0.6 * cm))

            # ── Schedule per team ────────────────────────────────────────────
            if not df_sched.empty:
                p_sched = df_sched[df_sched["Poule"] == poule]
                all_teams_in_poule = sorted(
                    set(p_sched["Team A"].tolist() + p_sched["Team B"].tolist())
                )

                story.append(Paragraph("Match Schedule", S_SECTION))
                story.append(Spacer(1, 0.2 * cm))

                sched_hdrs = ["Date", "Time", "Terrain", "Team", "Opponent"]
                sched_rows = []
                for team in all_teams_in_poule:
                    mask = (p_sched["Team A"] == team) | (p_sched["Team B"] == team)
                    team_matches = p_sched[mask].sort_values(["Date", "Time"])
                    for rn, (_, row) in enumerate(team_matches.iterrows(), 1):
                        opp = row["Team B"] if row["Team A"] == team else row["Team A"]
                        sched_rows.append([
                            row["Date"], row["Time"], row["Terrain"],
                            team, opp,
                        ])
                    if len(team_matches) > 0:
                        sched_rows.append(["", "", "", "", ""])  # blank separator row

                if sched_rows:
                    story.append(_schedule_table(sched_rows, sched_hdrs))

    doc = SimpleDocTemplate(
        path,
        pagesize=A4,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=MARGIN,
        bottomMargin=MARGIN,
        title=f"{club_name} — Competition Schedule",
        author=club_name,
    )
    doc.build(story)
    print(f"Saved PDF to:      {path}")


def print_schedule(df_sched: pd.DataFrame) -> None:
    """Pretty-print the schedule grouped by poule and date."""
    if df_sched.empty:
        print("No matches scheduled.")
        return
    has_poule = "Poule" in df_sched.columns and df_sched["Poule"].nunique() > 1
    poule_groups = df_sched.groupby("Poule") if has_poule else [(None, df_sched)]
    for poule_name, poule_df in poule_groups:
        if has_poule:
            print(f"\n══ Poule {poule_name} ════════════════════════════════════")
        for date, group in poule_df.groupby("Date"):
            print(f"── {date} ────────────────────────────────────")
            for _, row in group.iterrows():
                print(f"  {row['Time']}  [{row['Terrain']}]  {row['Team A']}  vs  {row['Team B']}")
    print()


def export_excel(
    df_sched: pd.DataFrame,
    df_unsched: pd.DataFrame,
    path: str,
    team_info_df: pd.DataFrame | None = None,
) -> None:
    """
    Export schedule to a colour-coded Excel workbook.

    Sheets
    -------
    - "Teams"        — team / player info (only when team_info_df is provided)
    - "Schedule"     — full schedule sorted by date / time / terrain
    - "By Team"      — per-team match list
    - "Poule X"      — one sheet per poule (when multiple poules exist)
    - "Unscheduled"  — matches that could not be placed (if any)
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        print("openpyxl not installed — saving as CSV instead.  pip install openpyxl")
        df_sched.to_csv(path.replace(".xlsx", "_schedule.csv"), index=False)
        return

    BLUE_HEADER  = PatternFill("solid", fgColor="2E75B6")
    GREEN_HEADER = PatternFill("solid", fgColor="375623")
    RED_HEADER   = PatternFill("solid", fgColor="C00000")
    ALT_ROW      = PatternFill("solid", fgColor="DEEAF1")
    WHITE_FONT   = Font(color="FFFFFF", bold=True)

    def _write_sheet(ws, df: pd.DataFrame, header_fill, col_widths=None):
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = WHITE_FONT
                elif r_idx % 2 == 0:
                    cell.fill = ALT_ROW
        ws.row_dimensions[1].height = 20
        for col in ws.columns:
            wanted = max((len(str(c.value or "")) for c in col), default=8) + 4
            ws.column_dimensions[col[0].column_letter].width = min(wanted, 30)

    wb = openpyxl.Workbook()

    # Sheet 1: team / player info (if available)
    ws_first = wb.active
    if team_info_df is not None and not team_info_df.empty:
        ws_first.title = "Teams"
        _write_sheet(ws_first, team_info_df, GREEN_HEADER)
        ws1 = wb.create_sheet("Schedule")
    else:
        ws_first.title = "Schedule"
        ws1 = ws_first

    # Schedule sheet
    _write_sheet(ws1, df_sched, BLUE_HEADER)

    # Sheet 2: per-team view
    if not df_sched.empty:
        by_team_rows = []
        all_teams = sorted(
            set(df_sched["Team A"].tolist() + df_sched["Team B"].tolist())
        )
        for team in all_teams:
            mask = (df_sched["Team A"] == team) | (df_sched["Team B"] == team)
            for _, row in df_sched[mask].iterrows():
                opponent = row["Team B"] if row["Team A"] == team else row["Team A"]
                by_team_rows.append(
                    {
                        "Poule": row.get("Poule", ""),
                        "Team": team,
                        "Date": row["Date"],
                        "Time": row["Time"],
                        "Terrain": row["Terrain"],
                        "Opponent": opponent,
                    }
                )
        df_by_team = pd.DataFrame(by_team_rows).sort_values(["Poule", "Team", "Date", "Time"])
        ws2 = wb.create_sheet("By Team")
        _write_sheet(ws2, df_by_team, BLUE_HEADER)

    # Sheet 3: one sheet per poule (if multiple poules)
    if not df_sched.empty and "Poule" in df_sched.columns and df_sched["Poule"].nunique() > 1:
        for poule_name in sorted(df_sched["Poule"].unique()):
            df_p = df_sched[df_sched["Poule"] == poule_name].reset_index(drop=True)
            ws_p = wb.create_sheet(f"Poule {poule_name}")
            _write_sheet(ws_p, df_p, BLUE_HEADER)

    # Sheet 4: unscheduled (optional)
    if not df_unsched.empty:
        ws3 = wb.create_sheet("Unscheduled")
        _write_sheet(ws3, df_unsched, RED_HEADER)

    wb.save(path)
    print(f"Saved schedule to: {path}")


# ── Static HTML export (GitHub Pages) ─────────────────────────────────────────



def export_html(
    df_sched: pd.DataFrame,
    team_info_df: "pd.DataFrame | None",
    path: str,
    club_name: str = "Tennis Club",
    season: str = "Season 2026",
) -> None:
    """Build the static HTML site and save schedule data to input/schedule.json."""
    import json
    from datetime import date as _date

    poules = (
        sorted(df_sched["Poule"].unique())
        if "Poule" in df_sched.columns
        else [""]
    )

    info_lkp: dict[str, dict] = {}
    if team_info_df is not None and not team_info_df.empty:
        for _, row in team_info_df.iterrows():
            name = str(row.get("Team", "") or "").strip()
            if name:
                info_lkp[name] = {
                    "player_1": str(row.get("Player 1", "") or "").strip(),
                    "player_2": str(row.get("Player 2", "") or "").strip(),
                    "tel_1":    str(row.get("Tel P1",   "") or "").strip(),
                    "tel_2":    str(row.get("Tel P2",   "") or "").strip(),
                }

    teams_by_poule: dict[str, list[dict]] = {}
    for poule in poules:
        df_p = (
            df_sched[df_sched["Poule"] == poule]
            if "Poule" in df_sched.columns and poule
            else df_sched
        )
        team_names = sorted(set(df_p["Team A"].tolist() + df_p["Team B"].tolist()))
        teams_by_poule[poule] = [
            {"name": t, **info_lkp.get(t, {})} for t in team_names
        ]

    matches = []
    for _, row in df_sched.iterrows():
        ta        = str(row["Team A"])
        tb        = str(row["Team B"])
        date_str  = str(row["Date"])
        time_str  = str(row["Time"])
        poule_str = str(row.get("Poule", "")) if "Poule" in row.index else ""
        terrain   = str(row["Terrain"])
        mid = (
            f"{date_str}_{time_str}_{ta}_{tb}"
            .replace(" ", "_").replace(".", "_")
            .replace("&", "and").replace("/", "-")
        )
        matches.append({
            "id":      mid,
            "poule":   poule_str,
            "date":    date_str,
            "time":    time_str,
            "terrain": terrain,
            "team_a":  ta,
            "team_b":  tb,
        })

    schedule_data = {
        "club_name":      club_name,
        "season":         season,
        "generated":      str(_date.today()),
        "poules":         list(poules),
        "teams_by_poule": teams_by_poule,
        "matches":        matches,
    }

    Path("input/schedule.json").write_text(
        json.dumps(schedule_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print("Saved schedule data to: input/schedule.json")

    site_builder.build(schedule_data, path)



# ── CLI ────────────────────────────────────────────────────────────────────────

def _parse_args():
    p = argparse.ArgumentParser(
        description="Poule competition scheduler using CP-SAT optimisation."
    )
    p.add_argument("--teams",   help="CSV with team names (column: Team)")
    p.add_argument("--avail",   help="CSV with team availabilities (Google Forms export)")
    p.add_argument("--slots",   help="CSV with terrain slot definitions")
    p.add_argument("--output",  default="schedule.xlsx", help="Output Excel file path")
    p.add_argument("--timelimit", type=int, default=60,
                   help="Solver time limit in seconds (default: 60)")
    return p.parse_args()


# ── Demo ───────────────────────────────────────────────────────────────────────

def _run_demo():
    """Built-in demo with 2 poules × 4 teams over 6 weeks."""
    print("Running built-in demo with 2 poules × 4 teams over 6 weeks...\n")

    # Venue: 6 Mondays, 3 time slots per evening, 4 terrains each
    dates = [
        "2026-05-04", "2026-05-11", "2026-05-18",
        "2026-05-25", "2026-06-01", "2026-06-08",
    ]
    times = ["18:00", "19:30", "21:00"]
    terrain_slots = generate_terrain_slots(
        dates, times, n_terrains=4,
        terrain_overrides={"2026-05-25": 2}   # only 2 terrains available that week
    )

    poules = {
        "A": ["Alpha", "Bravo", "Charlie", "Delta"],
        "B": ["Echo", "Foxtrot", "Golf", "Hotel"],
    }
    all_teams = [t for ts in poules.values() for t in ts]

    # Simulate availability: each team randomly available at ~60 % of slots
    all_slot_labels = list(dict.fromkeys(s["slot"] for s in terrain_slots))
    rng = random.Random(42)
    team_avail = {
        team: set(sl for sl in all_slot_labels if rng.random() < 0.60)
        for team in all_teams
    }

    print("Team availabilities (number of available slots):")
    for poule_name, teams in poules.items():
        print(f"  Poule {poule_name}:")
        for team in teams:
            print(f"    {team:10s}: {len(team_avail[team]):2d} / {len(all_slot_labels)} slots")
    print()

    scheduled, unscheduled = schedule(
        poules, team_avail, terrain_slots, time_limit_s=30
    )

    df_sched, df_unsched = to_dataframes(scheduled, unscheduled)
    print_schedule(df_sched)
    export_excel(df_sched, df_unsched, "schedule_demo.xlsx")
    export_schedule_overview_pdf(df_sched, "schedule_demo_overview.pdf",
                                 club_name="TC Kooike", season="Seizoen 2026")
    export_pdf(df_sched, None, "schedule_demo.pdf",
               club_name="TC Kooike", season="Seizoen 2026")
    export_html(df_sched, None, "schedule_demo.html",
                club_name="TC Kooike", season="Seizoen 2026")


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    args = _parse_args()

    # If no arguments given, run the demo
    if not any([args.teams, args.avail, args.slots]):
        _run_demo()
        return

    missing = [flag for flag, val in [("--teams", args.teams),
                                       ("--avail", args.avail),
                                       ("--slots", args.slots)] if not val]
    if missing:
        print(f"Error: missing required arguments: {', '.join(missing)}")
        print("Run without arguments for a built-in demo.")
        return

    teams_by_poule, team_info_df = load_teams(args.teams)
    all_teams      = [t for ts in teams_by_poule.values() for t in ts]
    team_avail     = load_team_availabilities(args.avail)
    terrain_sl     = load_terrain_slots(args.slots)

    n_poules = len(teams_by_poule)
    print(f"Loaded {n_poules} poule(s), {len(all_teams)} teams total, {len(terrain_sl)} terrain slots.")

    scheduled, unscheduled = schedule(
        teams_by_poule, team_avail, terrain_sl, time_limit_s=args.timelimit
    )

    df_sched, df_unsched = to_dataframes(scheduled, unscheduled)
    print_schedule(df_sched)
    export_excel(df_sched, df_unsched, args.output, team_info_df=team_info_df)
    pdf_path = args.output.replace(".xlsx", ".pdf")
    overview_pdf_path = args.output.replace(".xlsx", "_overview.pdf")
    html_path = args.output.replace(".xlsx", ".html")
    export_schedule_overview_pdf(df_sched, overview_pdf_path,
                                 club_name="TC Kooike", season="Seizoen 2026")
    export_pdf(df_sched, team_info_df, pdf_path,
               club_name="TC Kooike", season="Seizoen 2026")
    export_html(df_sched, team_info_df, html_path,
                club_name="TC Kooike", season="Seizoen 2026")


if __name__ == "__main__":
    main()
